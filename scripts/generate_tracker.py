"""
Generate project tracker Excel workbook.
Run: uv run python scripts/generate_tracker.py
Output: PROJECT_TRACKER.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import datetime

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "header_bg":    "1A1A2E",   # dark navy
    "header_fg":    "FFFFFF",
    "accent":       "4F8EF7",   # blue
    "done":         "22C55E",   # green
    "in_progress":  "F59E0B",   # amber
    "open":         "EF4444",   # red
    "blocked":      "8B5CF6",   # purple
    "low":          "94A3B8",   # slate
    "row_alt":      "F1F5F9",   # light grey
    "row_white":    "FFFFFF",
    "critical":     "FEE2E2",
    "high":         "FEF3C7",
    "medium":       "E0F2FE",
    "section_bg":   "1E3A5F",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header_row(ws, columns, row=1):
    for col_idx, (col_letter, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = fill(C["header_bg"])
        cell.font = font(bold=True, color=C["header_fg"], size=11)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def write_data_row(ws, row_idx, values, priority=None):
    bg = C["row_alt"] if row_idx % 2 == 0 else C["row_white"]
    if priority == "Critical": bg = C["critical"]
    elif priority == "High":   bg = C["high"]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.font = font(size=10)
        cell.alignment = left()
        cell.border = thin_border()
        # colour-code status cells
        if col_idx == values.__class__ and isinstance(values, (list, tuple)):
            pass
    # status column colouring
    status_cols = {v: i+1 for i, v in enumerate(values) if isinstance(v, str) and v in
                   ("Done", "Fixed", "Open", "In Progress", "Blocked", "Planned", "Partial")}
    status_map = {
        "Done": C["done"], "Fixed": C["done"],
        "In Progress": C["in_progress"], "Partial": C["in_progress"],
        "Open": C["open"], "Planned": C["open"],
        "Blocked": C["blocked"],
    }
    for val, col_i in status_cols.items():
        c = ws.cell(row=row_idx, column=col_i)
        c.font = font(bold=True, color=status_map.get(val, "000000"), size=10)

wb = Workbook()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "📊 Dashboard"
ws_dash.sheet_view.showGridLines = False
ws_dash.column_dimensions["A"].width = 28
ws_dash.column_dimensions["B"].width = 16
ws_dash.column_dimensions["C"].width = 28
ws_dash.column_dimensions["D"].width = 16

def dash_title(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(C["section_bg"])
    c.font = font(bold=True, color="FFFFFF", size=13)
    c.alignment = center()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 28

def dash_kv(ws, row, label, value, label2="", value2=""):
    for col, val, bg in [(1, label, "EFF6FF"), (2, value, "DBEAFE"),
                          (3, label2, "F0FDF4"), (4, value2, "DCFCE7")]:
        if val == "": continue
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=(col in (1, 3)), size=11)
        c.alignment = center()
        c.border = thin_border()

# Title
title = ws_dash.cell(row=1, column=1,
    value="VisionFlow — Project Tracker Dashboard")
title.fill = fill(C["header_bg"])
title.font = font(bold=True, color="FFFFFF", size=16)
title.alignment = center()
ws_dash.merge_cells("A1:D1")
ws_dash.row_dimensions[1].height = 36

ws_dash.cell(row=2, column=1,
    value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}").font = font(color="6B7280", size=9)
ws_dash.merge_cells("A2:D2")

dash_title(ws_dash, 4, "BUGS & ISSUES")
dash_kv(ws_dash, 5,  "Total Bugs",      "16",  "Fixed",       "10")
dash_kv(ws_dash, 6,  "Open",            "4",   "Blocked",     "1")
dash_kv(ws_dash, 7,  "Critical",        "4",   "High",        "6")

dash_title(ws_dash, 9, "FEATURES")
dash_kv(ws_dash, 10, "Total Features",  "18",  "Done",        "6")
dash_kv(ws_dash, 11, "In Progress",     "3",   "Planned",     "9")

dash_title(ws_dash, 13, "TECHNICAL DEBT")
dash_kv(ws_dash, 14, "Total Items",     "12",  "Critical",    "3")
dash_kv(ws_dash, 15, "High",            "5",   "Medium",      "4")

dash_title(ws_dash, 17, "QA & TESTING")
dash_kv(ws_dash, 18, "Total Tasks",     "21",  "Passing",     "11")
dash_kv(ws_dash, 19, "Missing Tests",   "10",  "Coverage ~",  "42%")

dash_title(ws_dash, 21, "COMPONENT HEALTH")
components = [
    ("Detection (YOLO + RF-DETR)",  "✅ Production Ready"),
    ("Tracking (ByteTrack)",         "✅ Production Ready"),
    ("Inference Pipeline",           "✅ Production Ready"),
    ("FastAPI Backend",              "✅ Production Ready"),
    ("Analytics / SQLite",           "✅ Production Ready"),
    ("Auto-Labeling Loop",           "✅ Production Ready"),
    ("Training + MLflow",            "✅ Production Ready"),
    ("DVC Reproducibility",          "✅ Configured"),
    ("Streamlit Dashboard",          "⚠️  Duplicate (deprecate?)"),
    ("Next.js Dashboard",            "✅ Production Ready"),
    ("Drift Detection",              "⚡ Partial (confidence only)"),
    ("Prometheus / Grafana",         "🔜 Planned"),
    ("Kafka Events",                 "⚡ Optional / Noop default"),
    ("GPU / ONNX Inference",         "🔜 Planned"),
    ("Docker Deployment",            "✅ Ready"),
    ("CI/CD GitHub Actions",         "⚡ Partial"),
]
for i, (comp, status) in enumerate(components, 22):
    c1 = ws_dash.cell(row=i, column=1, value=comp)
    c2 = ws_dash.cell(row=i, column=2, value=status)
    bg = "F0FDF4" if "✅" in status else "FEF3C7" if "⚡" in status else "FEE2E2" if "⚠️" in status else "EFF6FF"
    for c in (c1, c2):
        c.fill = fill(bg)
        c.font = font(size=10)
        c.alignment = left()
        c.border = thin_border()
    ws_dash.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — BUGS & ISSUES
# ══════════════════════════════════════════════════════════════════════════════
ws_bugs = wb.create_sheet("🐛 Bugs & Issues")
ws_bugs.sheet_view.showGridLines = False

bug_cols = [
    ("A", "ID",          7),
    ("B", "Title",       40),
    ("C", "Component",   20),
    ("D", "Priority",    12),
    ("E", "Status",      14),
    ("F", "Root Cause",  45),
    ("G", "Fix Applied", 45),
    ("H", "Notes",       30),
]
write_header_row(ws_bugs, bug_cols)

bugs = [
    ("B-01", "OpenCV MSMF initialization hangs on Windows",
     "API / Video Capture", "Critical", "Fixed",
     "cv2.VideoCapture() with CAP_MSMF backend hangs 5-30s or silently returns bad frames on Windows 10/11",
     "Replaced ThreadedVideoCapture with VidGear CamGear; passes backend=cv2.CAP_DSHOW via constructor arg",
     "DSHOW is deprecated on Win11; monitor for failures"),

    ("B-02", "VidGear CAP_PROP_BACKEND read-only error",
     "API / Video Capture", "Critical", "Fixed",
     "Passing CAP_PROP_BACKEND (id=42) via options dict triggers OpenCV error — it's a read-only prop, can't be set via cap.set()",
     "Pass backend as CamGear(backend=cv2.CAP_DSHOW) constructor arg instead of options dict",
     "VidGear 0.3.4 confirmed working"),

    ("B-03", "VidGear stream torn down on first None frame",
     "API / Video Capture", "Critical", "Fixed",
     "CamGear background thread hasn't captured first frame yet on initial read(); our code treated first None as fatal and destroyed stream",
     "Added _NULL_FRAME_THRESHOLD=20 grace period before tearing down stream",
     "20 reads × 10ms = 200ms warm-up tolerance"),

    ("B-04", "Recharts width(-1) height(-1) SSR warning",
     "Frontend / Dashboard", "High", "Fixed",
     "ResponsiveContainer runs on server during Next.js SSR; DOM not available → reports -1×-1 dimensions",
     "Added mounted state guard; use explicit height={160} instead of height='100%'",
     "Applied to TelemetryPanel.tsx"),

    ("B-05", "React hydration mismatch on VideoCanvas",
     "Frontend / Dashboard", "High", "Fixed",
     "useState(Date.now()) called at server render time produces different value than client hydration",
     "Initialize imgKey=0 (stable); useEffect sets real Date.now() after mount",
     ""),

    ("B-06", "Invalid Tailwind class pl-[-10px]",
     "Frontend / Dashboard", "Low", "Fixed",
     "Arbitrary negative padding not supported in Tailwind v3; class silently ignored",
     "Removed from CardContent className in TelemetryPanel",
     ""),

    ("B-07", "Sync generator in async FastAPI endpoint",
     "API / Streaming", "Critical", "Fixed",
     "feed_generator() was a blocking sync generator inside async def video_feed(); held event loop thread for full stream duration",
     "Converted to async generator; time.sleep → await asyncio.sleep; inference runs in loop.run_in_executor(None, ...) — event loop free during YOLO",
     ""),

    ("B-08", "Multiple /video_feed connections each spawn a VidGearCapture",
     "API / Streaming", "High", "Fixed",
     "Each HTTP connection instantiated a new VidGearCapture; DSHOW exclusive access → 2nd connection failed",
     "_acquire_capture()/_release_capture() with threading.Lock + ref-count; all connections share one singleton; cleanly released when last client disconnects",
     ""),

    ("B-09", "No error boundary around pipeline.process_frame()",
     "API / Inference", "Critical", "Fixed",
     "If YOLO OOMs or throws on a malformed frame, feed_generator() crashes and the stream dies permanently until server restart",
     "CircuitBreaker wraps all 6 pipeline stages; OPEN after N consecutive failures; stream degrades gracefully to raw frame; outer try/except in feed_generator as last-resort safety net",
     "SOLID-compliant: one CircuitBreaker class owns all failure/recovery logic"),

    ("B-10", "Global mutable state without locks",
     "API / Concurrency", "High", "Open",
     "current_stream_source, pipeline_state, current_telemetry are plain dicts mutated by concurrent request handlers; compound read-modify-write not atomic",
     "Use threading.Lock() or replace with a proper state object",
     "Low risk for single-user dev; real issue under load"),

    ("B-11", "/health endpoint returns hardcoded 'healthy'",
     "API / Reliability", "Medium", "Open",
     "Health check doesn't probe pipeline, model, or DB — will return healthy even if inference is broken",
     "Add pipeline is not None check + last_error timestamp + DB connectivity probe",
     ""),

    ("B-12", "Low-confidence frames dir grows unbounded",
     "Data / Storage", "Medium", "Open",
     "data/low_confidence_frames/ has no TTL or max-size limit; fills disk over time",
     "Add periodic cleanup: keep last N frames or frames younger than X days",
     ""),

    ("B-13", "SQLite single-writer contention",
     "Analytics / DB", "Low", "Open",
     "AnalyticsDB opens connection per call; concurrent API writes (analytics + labeling) can cause SQLite lock errors",
     "Use connection pool or serialize writes via queue",
     "Fine for single-user; matters under concurrent inference"),

    ("B-14", "Streamlit pages/1_inference.py bypasses FastAPI backend",
     "Frontend / Streamlit", "High", "Open",
     "Opens cv2.VideoCapture(0) directly; ignores VidGear, reconnection logic, pipeline_state toggles",
     "Either: (a) delete Streamlit inference page, (b) route it through the API",
     "Architectural confusion — two inference paths"),

    ("B-15", "No adaptive frame rate / frame dropping under load",
     "API / Performance", "Medium", "Open",
     "At 30fps input with 100ms inference latency, frames queue up faster than they're processed; no skip logic",
     "Drop frames if inference backlog exceeds threshold; or cap input to inference FPS",
     ""),

    ("B-16", "WebSocket telemetry broadcasts unconditionally every 500ms",
     "API / Performance", "Low", "Open",
     "Sends JSON to all connected clients even when no values changed",
     "Diff against previous telemetry; only broadcast on change",
     ""),
]

for i, bug in enumerate(bugs, 2):
    priority = bug[3]
    write_data_row(ws_bugs, i, bug, priority)
    ws_bugs.row_dimensions[i].height = 48

ws_bugs.freeze_panes = "A2"
ws_bugs.auto_filter.ref = f"A1:H{len(bugs)+1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — FEATURES
# ══════════════════════════════════════════════════════════════════════════════
ws_feat = wb.create_sheet("✨ Features")
ws_feat.sheet_view.showGridLines = False

feat_cols = [
    ("A", "ID",           7),
    ("B", "Feature",      38),
    ("C", "Area",         20),
    ("D", "Priority",     12),
    ("E", "Status",       14),
    ("F", "Description",  50),
    ("G", "Effort",       12),
    ("H", "Notes",        30),
]
write_header_row(ws_feat, feat_cols)

features = [
    # Done
    ("F-01", "YOLO11 detection with configurable precision",
     "Detection", "Critical", "Done",
     "YOLODetector supporting FP32/FP16; model cached via ModelRegistry singleton", "Done", ""),
    ("F-02", "RF-DETR secondary detector",
     "Detection", "High", "Done",
     "RFDETRDetector as fallback in DualDetector; COCO class ID mapping", "Done", ""),
    ("F-03", "DualDetector active learning loop",
     "Detection", "High", "Done",
     "3 modes: hot/inline/batch; saves low-confidence frames for offline analysis + retraining", "Done", ""),
    ("F-04", "ByteTrack multi-object tracking",
     "Tracking", "Critical", "Done",
     "Persistent visitor IDs across frames; supervision-compatible; version-aware API", "Done", ""),
    ("F-05", "MLflow experiment tracking",
     "MLOps", "High", "Done",
     "Full training run logging: params, metrics, artifacts, model registration; DagsHub remote", "Done", ""),
    ("F-06", "3-layer secrets management",
     "Infrastructure", "High", "Done",
     "YAML → ENV → Code injection; no credentials in git; works with Docker/K8s/CI", "Done", ""),
    ("F-07", "DVC data versioning pipeline",
     "MLOps", "High", "Done",
     "dvc.yaml: prepare_data → train → evaluate; reproducibility end-to-end", "Done", ""),
    ("F-08", "Triage UI (accept/reject/label frames)",
     "Active Learning", "High", "Done",
     "API endpoints + Next.js UI for reviewing low-confidence captures; moves to auto_labeled/", "Done", ""),
    ("F-09", "Prometheus metrics endpoint",
     "Monitoring", "Medium", "Done",
     "FastAPI Instrumentator exposes /metrics; ready for Grafana scrape", "Done", ""),
    ("F-10", "MJPEG live stream with VidGear",
     "Streaming", "Critical", "Done",
     "GET /video_feed; VidGearCapture with DSHOW backend, null-frame tolerance, exponential backoff", "Done", ""),
    # In Progress
    ("F-11", "Next.js production dashboard",
     "Frontend", "Critical", "In Progress",
     "Control panel, video canvas, telemetry panel, analytics charts, triage page; SSR hydration fixed", "Medium", "Active development"),
    ("F-12", "Confidence-based drift detection",
     "MLOps", "High", "In Progress",
     "DriftDetector sliding window; triggers retraining; confidence_drop method implemented", "Small", "Data drift (distribution shift) not yet done"),
    ("F-13", "GitHub Actions CI/CD",
     "Infrastructure", "Medium", "In Progress",
     "Automated test + lint pipeline; .github/workflows/ partially configured", "Medium", ""),
    # Planned
    ("F-14", "ONNX export for fast CPU inference",
     "Performance", "Critical", "Planned",
     "Export YOLO11n to ONNX; expected ~5× speedup over PyTorch CPU (100ms → 20ms per frame)", "Medium",
     "Biggest single performance win available"),
    ("F-15", "INT8 quantization",
     "Performance", "High", "Planned",
     "Post-training quantization; further 2× speedup; accuracy trade-off analysis needed", "Medium",
     "See docs/guides/MODEL_QUANTIZATION_GUIDE.md"),
    ("F-16", "GPU / CUDA inference support",
     "Performance", "High", "Planned",
     "Docker GPU support commented out in docker-compose; needs CUDA device selection logic", "Small", ""),
    ("F-17", "Distribution-shift drift detection",
     "MLOps", "Medium", "Planned",
     "Detect input data drift (color histogram, embedding distance) vs confidence-only", "Large", ""),
    ("F-18", "Grafana monitoring dashboard",
     "Monitoring", "Medium", "Planned",
     "Connect Prometheus /metrics to Grafana; panels for FPS, latency, object count, drift score", "Medium", ""),
    ("F-19", "Automated drift-triggered retraining",
     "MLOps", "High", "Planned",
     "DriftDetector → auto-trigger train.py → MLflow run → model promotion; currently manual only", "Medium", ""),
    ("F-20", "Live demo on HuggingFace Spaces",
     "Distribution", "High", "Planned",
     "Deploy with sample video (no webcam required); Docker-based; public visibility", "Medium",
     "Highest ROI for project visibility"),
    ("F-21", "Multi-camera support",
     "Streaming", "Medium", "Planned",
     "Multiple concurrent VidGearCapture instances; stream selection in UI", "Large", ""),
    ("F-22", "Kafka event streaming",
     "Infrastructure", "Low", "Planned",
     "KafkaPipelineEventPublisher already stubbed; needs broker config + consumer example", "Small", ""),
]

for i, feat in enumerate(features, 2):
    priority = feat[3]
    write_data_row(ws_feat, i, feat, priority)
    ws_feat.row_dimensions[i].height = 48

ws_feat.freeze_panes = "A2"
ws_feat.auto_filter.ref = f"A1:H{len(features)+1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — TECHNICAL DEBT
# ══════════════════════════════════════════════════════════════════════════════
ws_debt = wb.create_sheet("⚙️ Tech Debt")
ws_debt.sheet_view.showGridLines = False

debt_cols = [
    ("A", "ID",           7),
    ("B", "Issue",        40),
    ("C", "Area",         20),
    ("D", "Priority",     12),
    ("E", "Status",       14),
    ("F", "Impact",       40),
    ("G", "Recommended Fix", 45),
    ("H", "Effort",       10),
]
write_header_row(ws_debt, debt_cols)

debt = [
    ("D-01", "Two frontends: Streamlit AND Next.js",
     "Architecture", "Critical", "Open",
     "Maintenance burden; contributor confusion; Streamlit inference page bypasses the entire API",
     "Decision needed: deprecate Streamlit or keep as internal debug-only tool; remove from README",
     "Medium"),
    ("D-02", "Sync generator in async endpoint (feed_generator)",
     "API / Performance", "Critical", "Open",
     "Blocks one uvicorn worker thread per active stream; no concurrency possible on that worker",
     "asyncio.to_thread(feed_generator) or rewrite as async generator",
     "Small"),
    ("D-03", "One VidGearCapture per /video_feed connection",
     "API / Resources", "Critical", "Open",
     "Two browser tabs → two camera handles → DSHOW exclusive access failure",
     "Singleton VidGearCapture at module level; share across all connections; ref-count releases",
     "Small"),
    ("D-04", "Global mutable dict state (pipeline_state, current_telemetry)",
     "API / Concurrency", "High", "Open",
     "Not thread-safe under concurrent requests; read-modify-write race on pipeline toggles",
     "Replace with a proper state class with threading.RLock; or use asyncio primitives",
     "Small"),
    ("D-05", "ModelRegistry has no eviction",
     "Memory", "Medium", "Open",
     "Models cached indefinitely in class dict; loading multiple model variants fills RAM",
     "Add LRU eviction or max-models limit",
     "Small"),
    ("D-06", "SQLite no connection pool",
     "Analytics / DB", "Medium", "Open",
     "AnalyticsDB opens connection per method call; lock contention under concurrent API calls",
     "Use sqlite3 with check_same_thread=False + threading.Lock, or switch to aiosqlite",
     "Small"),
    ("D-07", "No frame rate cap / backlog protection in feed_generator",
     "Performance", "High", "Open",
     "Inference at 10 FPS with 30 FPS input causes growing frame backlog and memory pressure",
     "Add frame_skip_counter; only process every Nth frame if inference is slow",
     "Small"),
    ("D-08", "MJPEG streaming instead of binary WebSocket frames",
     "Performance", "Medium", "Open",
     "MJPEG sends full JPEG headers per frame; less efficient than binary WebSocket frames",
     "Migrate /video_feed to WebSocket binary (or HLS for multi-viewer); MJPEG fine for MVP",
     "Large"),
    ("D-09", "Streamlit pages duplicate analytics queries already in API",
     "Code Duplication", "Medium", "Open",
     "pages/3_analytics.py queries SQLite directly; duplicates /analytics/stats API logic",
     "Streamlit pages should call the API, not the DB directly",
     "Medium"),
    ("D-10", "YOLO model path hardcoded as 'yolo11n' (no .pt extension fallback)",
     "Config", "Low", "Open",
     "Config says model.name: yolo11n; relies on Ultralytics auto-download; no local path fallback",
     "Support explicit model path in config; add exists check before download",
     "Small"),
    ("D-11", "DriftDetector only supports confidence_drop method",
     "MLOps", "Medium", "Open",
     "drift.method config accepts other values but only confidence_drop is implemented",
     "Add distribution shift method or raise NotImplementedError for unsupported methods",
     "Medium"),
    ("D-12", "No request/response logging middleware",
     "Observability", "Low", "Open",
     "Errors and slow requests not logged with context (path, duration, status)",
     "Add FastAPI middleware for structured request logging",
     "Small"),
]

for i, d in enumerate(debt, 2):
    priority = d[3]
    write_data_row(ws_debt, i, d, priority)
    ws_debt.row_dimensions[i].height = 48

ws_debt.freeze_panes = "A2"
ws_debt.auto_filter.ref = f"A1:H{len(debt)+1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — QA & TESTING
# ══════════════════════════════════════════════════════════════════════════════
ws_qa = wb.create_sheet("🧪 QA & Testing")
ws_qa.sheet_view.showGridLines = False

qa_cols = [
    ("A", "ID",           7),
    ("B", "Test / Task",  42),
    ("C", "Type",         18),
    ("D", "Priority",     12),
    ("E", "Status",       14),
    ("F", "Description",  48),
    ("G", "File",         35),
]
write_header_row(ws_qa, qa_cols)

qa = [
    # Existing passing tests
    ("Q-01", "Config load / merge / validate",
     "Unit", "High", "Done",
     "Tests for load_config, merge_configs, validate_config, inject_secrets",
     "tests/test_config.py"),
    ("Q-02", "DetectorFactory singleton caching",
     "Unit", "High", "Done",
     "Ensures same model instance returned for same config; different configs return different instances",
     "tests/test_detector_factory.py"),
    ("Q-03", "DualDetector modes (hot/inline/batch) + frame saving",
     "Unit", "High", "Done",
     "Covers all 3 dual-detector modes; frame save count; secondary_ratio tracking",
     "tests/test_dual_detector.py"),
    ("Q-04", "TrackerFactory non-singleton behavior",
     "Unit", "High", "Done",
     "Confirms trackers are always new instances (stateful); ByteTrack/BoTSORT/OC-SORT",
     "tests/test_tracker_factory.py"),
    ("Q-05", "InferencePipeline process_frame + run_offline",
     "Integration", "Critical", "Done",
     "End-to-end frame processing; analytics accumulation; offline video file inference",
     "tests/test_pipeline.py"),
    ("Q-06", "VisitorAnalytics dwell time",
     "Unit", "High", "Done",
     "Unique visitor tracking; dwell time computation per tracker_id",
     "tests/test_visitor_analytics.py"),
    ("Q-07", "AnalyticsDB save / fetch",
     "Unit", "High", "Done",
     "SQLite persistence: save_inference_run, get_analytics_summary, visitor_analytics",
     "tests/test_analytics.py"),
    ("Q-08", "DriftDetector confidence window",
     "Unit", "Medium", "Done",
     "Sliding window; drift detection threshold; get_metrics()",
     "tests/test_drift.py"),
    ("Q-09", "AutoLabeler local export",
     "Unit", "Medium", "Done",
     "Label collection from DualDetector; YOLO format conversion; local export",
     "tests/test_auto_labeler.py"),
    ("Q-10", "FastAPI endpoint smoke tests",
     "Integration", "High", "Done",
     "GET /health, POST /predict/video, GET /config — basic response code checks",
     "tests/test_api.py"),
    # Missing / needed
    ("Q-11", "VidGearCapture null-frame grace period",
     "Unit", "Critical", "Open",
     "Confirm stream not torn down before _NULL_FRAME_THRESHOLD consecutive nulls; mock CamGear",
     "tests/test_video_capture.py (create)"),
    ("Q-12", "Concurrent /video_feed connections",
     "Integration", "High", "Open",
     "Open 2 connections; confirm single VidGearCapture used (once singleton fix applied)",
     "tests/test_api.py"),
    ("Q-13", "feed_generator error recovery",
     "Integration", "Critical", "Open",
     "Inject exception into process_frame(); confirm stream continues with raw frame passthrough",
     "tests/test_api.py"),
    ("Q-14", "ONNX model inference parity",
     "Performance", "High", "Open",
     "Compare ONNX vs PyTorch detections on same frame; assert mAP delta < threshold",
     "tests/test_onnx.py (create)"),
    ("Q-15", "Training pipeline end-to-end",
     "Integration", "High", "Open",
     "Run scripts/train.py with coco8 dataset; confirm MLflow run created, metrics.json saved",
     "tests/test_training_pipeline.py (create)"),
    ("Q-16", "DVC pipeline reproducibility",
     "Integration", "Medium", "Open",
     "Run dvc repro; confirm outputs match dvc.lock; no stale stages",
     "tests/test_dvc.py (create)"),
    ("Q-17", "Drift-triggered retraining flow",
     "Integration", "Medium", "Open",
     "Inject low-confidence stream; confirm drift detected; confirm training event created",
     "tests/test_drift_trigger.py (create)"),
    ("Q-18", "Frontend E2E: monitor page loads + stream visible",
     "E2E", "Medium", "Open",
     "Playwright: navigate to /; confirm video feed renders; WebSocket connects",
     "frontend/tests/e2e/ (create)"),
    ("Q-19", "Frontend E2E: triage page accept/reject flow",
     "E2E", "Medium", "Open",
     "Playwright: confirm low-conf frame appears; accept; confirm moved to auto_labeled",
     "frontend/tests/e2e/ (create)"),
    ("Q-20", "Performance benchmark: FPS at various models",
     "Performance", "High", "Open",
     "Measure end-to-end FPS: YOLO11n PyTorch vs ONNX vs INT8; document in docs/guides/SCALING.md",
     "scripts/benchmark.py (create)"),
    ("Q-21", "CircuitBreaker unit tests",
     "Unit", "Critical", "Done",
     "CLOSED→OPEN transition; OPEN fast-fail without calling fn; OPEN→HALF_OPEN→CLOSED recovery; half-open re-trip; kwargs forwarding",
     "tests/test_circuit_breaker.py"),
]

for i, q in enumerate(qa, 2):
    priority = q[3]
    write_data_row(ws_qa, i, q, priority)
    ws_qa.row_dimensions[i].height = 48

ws_qa.freeze_panes = "A2"
ws_qa.auto_filter.ref = f"A1:G{len(qa)+1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — MLOPS TASKS
# ══════════════════════════════════════════════════════════════════════════════
ws_mlops = wb.create_sheet("🚀 MLOps")
ws_mlops.sheet_view.showGridLines = False

mlops_cols = [
    ("A", "ID",           7),
    ("B", "Task",         40),
    ("C", "Area",         20),
    ("D", "Priority",     12),
    ("E", "Status",       14),
    ("F", "Description",  50),
    ("G", "Effort",       10),
    ("H", "Reference",    30),
]
write_header_row(ws_mlops, mlops_cols)

mlops = [
    ("M-01", "MLflow tracking — fully integrated",
     "Experiment Tracking", "Critical", "Done",
     "MLflowCallback logs params/metrics/artifacts; DagsHub remote URI; model registration",
     "Done", "src/vision_ml/training/callbacks.py"),
    ("M-02", "DVC pipeline configured (prepare→train→evaluate)",
     "Reproducibility", "High", "Done",
     "dvc.yaml with 3 stages; dependency tracking on configs; dvc.lock in repo",
     "Done", "dvc.yaml"),
    ("M-03", "Multi-source data ingestion with validation",
     "Data Pipeline", "High", "Done",
     "scripts/prepare_data.py: local + Roboflow; label schema validation; priority deduplication",
     "Done", "scripts/prepare_data.py"),
    ("M-04", "YOLO training with MLflow",
     "Training", "Critical", "Done",
     "Trainer._run_training(); manual + drift-triggered; event publishing; AnalyticsDB logging",
     "Done", "src/vision_ml/training/trainer.py"),
    ("M-05", "Confidence-based drift detection",
     "Drift Monitoring", "High", "In Progress",
     "DriftDetector sliding window implemented; auto-trigger not yet wired to training",
     "Partial", "src/vision_ml/training/drift_detector.py"),
    ("M-06", "Wire DVC to actual remote storage",
     "Reproducibility", "High", "Open",
     "DVC configured but .dvc/config remote not set up with real S3/DagsHub credentials",
     "Small", "docs/guides/DAGSHUB_MLFLOW_DVC_INTEGRATION.md"),
    ("M-07", "Auto drift-triggered retraining pipeline",
     "Automation", "High", "Open",
     "DriftDetector → subprocess or async call to train.py → new MLflow run → model promotion",
     "Medium", ""),
    ("M-08", "Model promotion workflow",
     "Model Registry", "Medium", "Open",
     "Define staging→production promotion criteria; MLflow model aliases; rollback procedure",
     "Medium", "scripts/model_registry_cli.py"),
    ("M-09", "ONNX export + benchmark",
     "Performance", "Critical", "Open",
     "Export best.pt → model.onnx; benchmark FPS/latency vs PyTorch; update SCALING.md",
     "Medium", "docs/guides/MODEL_QUANTIZATION_GUIDE.md"),
    ("M-10", "INT8 quantization + trade-off analysis",
     "Performance", "High", "Open",
     "Post-training quantization; measure mAP drop vs latency gain; document findings",
     "Medium", "docs/guides/MODEL_QUANTIZATION_GUIDE.md"),
    ("M-11", "Data drift (distribution shift) detection",
     "Drift Monitoring", "Medium", "Open",
     "Histogram comparison or embedding distance on input frames; beyond confidence-only",
     "Large", ""),
    ("M-12", "Grafana dashboard connected to Prometheus",
     "Monitoring", "Medium", "Open",
     "Prometheus /metrics is live; need Grafana datasource + panels (FPS, latency, objects, drift)",
     "Medium", ""),
    ("M-13", "CI/CD: automated test + lint on PR",
     "Infrastructure", "High", "In Progress",
     ".github/workflows/ exists; needs complete pytest + ruff workflow",
     "Small", ".github/workflows/"),
    ("M-14", "CI/CD: Docker build + push on merge to main",
     "Infrastructure", "Medium", "Open",
     "Build and push API + frontend images to registry on successful main merge",
     "Small", ""),
    ("M-15", "Scheduled retraining (cron-based)",
     "Automation", "Low", "Open",
     "train.py supports --trigger daily/weekly; needs cron job or Airflow DAG",
     "Medium", "config/training/base.yaml → schedule.mode"),
    ("M-16", "Scalability benchmarks (1→10→100 streams)",
     "Performance", "Medium", "Open",
     "Document CPU/GPU/memory at 1, 10, 100 concurrent streams; update SCALING.md",
     "Large", "docs/guides/SCALING.md"),
]

for i, m in enumerate(mlops, 2):
    priority = m[3]
    write_data_row(ws_mlops, i, m, priority)
    ws_mlops.row_dimensions[i].height = 48

ws_mlops.freeze_panes = "A2"
ws_mlops.auto_filter.ref = f"A1:H{len(mlops)+1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — COMPONENTS INVENTORY
# ══════════════════════════════════════════════════════════════════════════════
ws_comp = wb.create_sheet("📦 Components")
ws_comp.sheet_view.showGridLines = False

comp_cols = [
    ("A", "Module",         28),
    ("B", "File(s)",        40),
    ("C", "What It Does",   55),
    ("D", "Status",         18),
    ("E", "Test Coverage",  15),
    ("F", "Notes",          35),
]
write_header_row(ws_comp, comp_cols)

components_inv = [
    # Detection
    ("BaseDetector (ABC)",
     "src/vision_ml/detection/base.py",
     "Abstract interface for all detectors: load_model(), detect(), detect_batch()",
     "Done", "✅ Via factory tests", ""),
    ("YOLODetector",
     "src/vision_ml/detection/yolo_detector.py",
     "YOLO11 wrapper; FP32/FP16 precision; confidence/IOU configurable; model cached in ~/.cache/yolo/",
     "Done", "✅ test_detection.py", ""),
    ("RFDETRDetector",
     "src/vision_ml/detection/rfdetr_detector.py",
     "RF-DETR wrapper; COCO class ID mapping; returns supervision.Detections",
     "Done", "⚡ Minimal", "rf-detr-base.pth = 372MB"),
    ("DualDetector",
     "src/vision_ml/detection/dual_detector.py",
     "Ensemble: YOLO primary + RF-DETR fallback; 3 modes (hot/inline/batch); tracks secondary_ratio",
     "Done", "✅ test_dual_detector.py", "Core active learning component"),
    ("DetectorFactory + ModelRegistry",
     "src/vision_ml/detection/detector_factory.py\nsrc/vision_ml/detection/model_registry.py",
     "Factory pattern with singleton model caching; supports yolo11n/s/m/l/x + rfdetr",
     "Done", "✅ test_detector_factory.py", "No eviction strategy (D-05)"),
    # Tracking
    ("BaseTracker (ABC)",
     "src/vision_ml/tracking/base.py",
     "Abstract interface: update(), reset()",
     "Done", "✅ Via factory tests", ""),
    ("ByteTrackTracker",
     "src/vision_ml/tracking/bytetrack.py",
     "ByteTrack wrapper via GenericSVTracker; always new instances (stateful); version-aware params",
     "Done", "✅ test_tracker_factory.py", "Supports BoTSORT/OC-SORT too"),
    # Inference
    ("InferencePipeline",
     "src/vision_ml/inference/pipeline.py",
     "Main orchestrator: detection→tracking→analytics→labeling→drift; online + offline modes; process_frame()",
     "Done", "✅ test_pipeline.py", "234 lines; core production component"),
    # Analytics
    ("VisitorAnalytics",
     "src/vision_ml/analytics/visitor_analytics.py",
     "Tracks unique visitors by tracker_id; dwell time per person (first/last frame, duration)",
     "Done", "✅ test_visitor_analytics.py", ""),
    ("AnalyticsDB",
     "src/vision_ml/analytics/analytics_db.py",
     "SQLite persistence: inference_runs, visitor_analytics, training_events, labeling_events",
     "Done", "✅ test_analytics.py", "No connection pool (D-06)"),
    # Annotation
    ("FrameAnnotator",
     "src/vision_ml/annotation/annotator.py",
     "Supervision-based bbox + label + trace annotation; version-compatible; build_labels() formats display",
     "Done", "✅ test_annotator.py", ""),
    # Labeling
    ("AutoLabeler",
     "src/vision_ml/labeling/auto_labeler.py",
     "Collects low-conf frames from DualDetector; YOLO format export; optional Roboflow upload",
     "Done", "✅ test_auto_labeler.py", ""),
    # Training
    ("Trainer",
     "src/vision_ml/training/trainer.py",
     "YOLO training via Ultralytics; manual + drift-triggered; MLflow callback; saves metrics.json",
     "Done", "⚡ Partial", ""),
    ("MLflowCallback",
     "src/vision_ml/training/callbacks.py",
     "Experiment tracking: run management, param/metric logging, model registration, NaN sanitization",
     "Done", "⚡ Partial", "DagsHub URI parsing included"),
    ("DriftDetector",
     "src/vision_ml/training/drift_detector.py",
     "Confidence sliding window; drift score; check_interval; get_metrics(); triggers retraining",
     "Partial", "✅ test_drift.py", "Only confidence_drop method implemented"),
    # API
    ("FastAPI App",
     "src/vision_ml/api/main.py",
     "16 endpoints; CORS; Prometheus; VidGearCapture; feed_generator (MJPEG); WebSocket telemetry",
     "Done", "⚡ Smoke tests only", "Sync generator issue (D-02)"),
    ("VidGearCapture",
     "src/vision_ml/api/main.py",
     "VidGear CamGear wrapper; DSHOW on Windows; null-frame tolerance; exponential backoff reconnect",
     "Done", "❌ No tests", "Replaced ThreadedVideoCapture"),
    # Utils
    ("CircuitBreaker",
     "src/vision_ml/utils/circuit_breaker.py",
     "CLOSED→OPEN→HALF_OPEN state machine; wraps any callable; failure_threshold + recovery_frames configurable; used by InferencePipeline for all 6 stages",
     "Done", "✅ test_circuit_breaker.py", ""),
    ("Config System",
     "src/vision_ml/utils/config.py",
     "load_config, inject_secrets, merge_configs, validate_config; ENV var injection",
     "Done", "✅ test_config.py", ""),
    ("Logger",
     "src/vision_ml/logging/logger.py",
     "Centralized logging factory; ISO timestamps; console + optional file handler",
     "Done", "✅ Implicit", ""),
    # Events
    ("Event System",
     "src/vision_ml/events/base.py\nsrc/vision_ml/events/publishers.py",
     "EventType enum, Event dataclass, JobState; NoopPublisher (default) + KafkaPublisher (optional)",
     "Done", "❌ No tests", "Noop by default; Kafka optional"),
    # Frontend - Next.js
    ("VideoCanvas",
     "frontend/src/components/monitor/VideoCanvas.tsx",
     "MJPEG stream display; WebSocket telemetry receiver; auto-reconnect every 3s",
     "Done", "❌ No E2E", "Hydration fix applied"),
    ("TelemetryPanel",
     "frontend/src/components/monitor/TelemetryPanel.tsx",
     "FPS/latency/objects cards; Recharts confidence drift chart; live event log",
     "Done", "❌ No E2E", "SSR mount guard applied"),
    ("ControlPanel",
     "frontend/src/components/monitor/ControlPanel.tsx",
     "Stream source switcher; detection/tracking/annotation toggles; confidence/IoU sliders",
     "Done", "❌ No E2E", ""),
    ("Analytics Pages",
     "frontend/src/app/analytics/",
     "DriftChart, VisitorHeatmap, ClassDistribution, DataTable; pulls from /analytics/* API",
     "Done", "❌ No E2E", ""),
    ("Triage Page",
     "frontend/src/app/triage/",
     "Displays low-confidence frames; accept/reject/label actions via /triage/* API",
     "Done", "❌ No E2E", ""),
    # Streamlit
    ("Streamlit Home",
     "home.py",
     "Landing page; system status; summary metrics from AnalyticsDB; links to all pages",
     "Done", "N/A", "Duplicate UI — decision pending"),
    ("Streamlit Inference",
     "pages/1_inference.py",
     "Detector mode, video upload/webcam; direct cv2.VideoCapture (bypasses API)",
     "Done", "N/A", "Bypasses FastAPI (B-14)"),
    ("Streamlit Analytics",
     "pages/3_analytics.py",
     "Visitor metrics, dwell time charts, inference run table; queries SQLite directly",
     "Done", "N/A", "Duplicates API queries (D-09)"),
    ("Streamlit Training",
     "pages/4_training.py\npages/5_mlflow_experiments.py\npages/7_training_pipeline.py",
     "Training control; MLflow experiment browser; DVC pipeline status; event log",
     "Done", "N/A", ""),
]

for i, c in enumerate(components_inv, 2):
    bg = C["row_alt"] if i % 2 == 0 else C["row_white"]
    for col_idx, val in enumerate(c, 1):
        cell = ws_comp.cell(row=i, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.font = font(size=10)
        cell.alignment = left()
        cell.border = thin_border()
    ws_comp.row_dimensions[i].height = 52

ws_comp.freeze_panes = "A2"
ws_comp.auto_filter.ref = f"A1:F{len(components_inv)+1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — VISIBILITY & GROWTH
# ══════════════════════════════════════════════════════════════════════════════
ws_vis = wb.create_sheet("📣 Visibility")
ws_vis.sheet_view.showGridLines = False

vis_cols = [
    ("A", "ID",          7),
    ("B", "Action",      38),
    ("C", "Channel",     20),
    ("D", "Priority",    12),
    ("E", "Status",      14),
    ("F", "Detail / Why", 50),
    ("G", "Effort",      10),
]
write_header_row(ws_vis, vis_cols)

visibility = [
    ("V-01", "Add GIF/screenshot of dashboard to README top",
     "GitHub", "Critical", "Open",
     "First thing visitors see; no GIF = no clicks; record 60s screen capture of live dashboard",
     "1h"),
    ("V-02", "Deploy live demo on HuggingFace Spaces",
     "Distribution", "Critical", "Open",
     "Use sample video (not webcam); no setup required for visitors; massive discoverability boost",
     "3h"),
    ("V-03", "Add GitHub topics/tags",
     "GitHub", "High", "Open",
     "Add: computer-vision, yolo, bytetrack, mlops, fastapi, nextjs, object-tracking, retail-analytics",
     "5min"),
    ("V-04", "Reposition README: forkable template angle",
     "GitHub", "High", "Open",
     "'Fork this, swap the model, ship any CV system' — much broader appeal than retail-specific pitch",
     "1h"),
    ("V-05", "Post on Hacker News (Show HN)",
     "Community", "High", "Open",
     "Highest ROI single post; submit on a Tuesday/Wednesday morning; title: Show HN: [project]",
     "30min"),
    ("V-06", "Tweet tagging @Ultralytics and @Roboflow",
     "Social", "High", "Open",
     "Both actively amplify community projects; include demo GIF; tag on same day as HN post",
     "30min"),
    ("V-07", "Write technical blog post (Dev.to / Hashnode)",
     "Content", "Medium", "Open",
     "'Building a production CV pipeline: YOLO11 → ByteTrack → FastAPI → Next.js'; focus on decisions made",
     "4h"),
    ("V-08", "Submit to awesome-mlops and awesome-computer-vision lists",
     "Distribution", "Medium", "Open",
     "Passive, permanent traffic; open PRs to both lists with 1-line description",
     "1h"),
    ("V-09", "Post to r/MachineLearning and r/computervision",
     "Community", "Medium", "Open",
     "Targeted audience; screenshot + short description; link to HF Spaces demo",
     "30min"),
    ("V-10", "Add benchmark numbers to README",
     "GitHub", "Medium", "Open",
     "FPS at PyTorch vs ONNX vs INT8; real numbers build credibility",
     "2h"),
    ("V-11", "Pin repo on GitHub profile",
     "GitHub", "Low", "Open",
     "Ensure this is one of the 6 pinned repos on your profile",
     "5min"),
]

for i, v in enumerate(visibility, 2):
    priority = v[3]
    write_data_row(ws_vis, i, v, priority)
    ws_vis.row_dimensions[i].height = 42

ws_vis.freeze_panes = "A2"
ws_vis.auto_filter.ref = f"A1:G{len(visibility)+1}"


# ── Save ──────────────────────────────────────────────────────────────────────
output_path = "PROJECT_TRACKER.xlsx"
wb.save(output_path)
sheet_names = [ws.title.encode("ascii", "ignore").decode() for ws in wb.worksheets]
print(f"Saved: {output_path}  ({len(sheet_names)} sheets)")
